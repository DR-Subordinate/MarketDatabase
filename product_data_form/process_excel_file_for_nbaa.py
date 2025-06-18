from openpyxl import load_workbook

def process_excel(file_path):
    """
    Process Excel file
    Return list of item dictionaries
    """
    wb = load_workbook(file_path, data_only=True)
    ws = wb["出品リスト１"]

    products = []

    # Find all valid box numbers by scanning column B for "箱番"
    box_positions = []

    # First, collect all box positions
    for row_num in range(1, ws.max_row + 1):
        if ws[f"B{row_num}"].value == "箱番":
            box_number = ws[f"C{row_num}"].value
            if box_number and isinstance(box_number, int):
                box_positions.append((row_num, box_number))

    # Process each box section separately
    for i, (box_row, box_number) in enumerate(box_positions):
        # Determine the end row for this box section
        start_row = box_row + 4
        if i + 1 < len(box_positions):
            # If there's another box, stop before the next box
            end_row = box_positions[i + 1][0] - 1
        else:
            # If this is the last box, process to the end
            end_row = ws.max_row + 1

        # Process products for this box
        for data_row in range(start_row, end_row):
            # Get brand_name first and check if it exists
            brand_name = ws[f"B{data_row}"].value or ""
            brand_name = str(brand_name).strip() if brand_name else ""

            # Skip this row entirely if no brand name
            if not brand_name:
                continue

            # Only then check item number and get other data
            item_no = ws[f"A{data_row}"].value

            if item_no and isinstance(item_no, int):
                # Get values
                name = ws[f"C{data_row}"].value or ""
                line_material = ws[f"D{data_row}"].value or ""
                model_number = ws[f"E{data_row}"].value or ""
                color = ws[f"F{data_row}"].value or ""
                imprint = ws[f"G{data_row}"].value or ""
                condition = ws[f"H{data_row}"].value or ""
                cost_price = ws[f"K{data_row}"].value or ""
                winning_bid = ws[f"L{data_row}"].value or ""

                # Clean strings
                name = str(name).strip() if name else ""
                line_material = str(line_material).strip() if line_material else ""
                model_number = str(model_number).strip() if model_number else ""
                color = str(color).strip() if color else ""
                imprint = str(imprint).strip() if imprint else ""
                condition = str(condition).strip() if condition else ""

                # Concatenate 箱番 and NO
                number = f"{box_number}-{item_no}"  # 30-1, 76-1, etc.

                # Concatenate ライン・素材 and カラー
                material_color = f"{line_material} {color}"

                product = {
                    "number": number,
                    "brand_name": brand_name,
                    "name": name,
                    "material_color": material_color,
                    "model_number": model_number,
                    "imprint": imprint,
                    "condition": condition,
                    "cost_price": cost_price,
                    "winning_bid": winning_bid,
                }

                products.append(product)

    return products

if __name__ == "__main__":
    # Add file path
    file_path = ""

    products = process_excel(file_path)

    print(f"Processed {len(products)} products")

    for i, product in enumerate(products):
        print(f"\nProduct {i+1}:")
        for key, value in product.items():
            print(f"  {key}: {value}")
